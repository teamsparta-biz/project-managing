#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_rows 결과를 「인건비 정산 요청 시트」와 같은 포맷의 xlsx 로 내보낸다.

헤더 문구·열 순서는 원본 워크북의 '인건비 정산 요청 시트' 1행을 그대로 읽어 쓴다
(문구가 바뀌어도 따라간다). 값만 채우고 수식은 넣지 않는다.

usage:
  python export_xlsx.py --month 2026-07 --pm 송찬호 --out "산출물/인건비/2026-07.xlsx"
"""
import argparse
import datetime as dt
import io
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import build_rows as br

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", newline="\n")

SHEET = "인건비 정산 요청 시트"
DATE_COLS = {"H", "I"}
NUM_COLS = {"R"}
# 원본 시트의 대략적인 열 너비. 붙여넣기 전 눈으로 검수하기 좋은 정도로만 맞춘다.
WIDTHS = {
    "A": 22, "B": 9, "C": 12, "D": 26, "E": 10, "F": 10, "G": 10,
    "H": 12, "I": 12, "J": 11, "K": 10, "L": 20, "M": 26, "N": 16,
    "O": 8, "P": 18, "Q": 46, "R": 14, "S": 18, "T": 16, "U": 12,
}


ROLE_FILL = {
    "주강사": "DDEBF7",
    "기술튜터": "FFF2CC",
}


VERDICT_FILL = {
    "제출<계산": "F8CBAD",   # 과소 제출 의심
    "미제출": "F8CBAD",
    "제출>계산": "FFE699",   # 경비·초과근무일 수 있음
    "배정없음": "FFE699",
    "단가미정": "D9D9D9",
    "일치": "E2EFDA",
}


def _style_header(ws, widths, height=32):
    fill = PatternFill("solid", fgColor="D9D9D9")
    for i in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, size=9)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = height
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(ws.max_column), ws.max_row)


def _sheet_axhub(wb, assigns):
    """ax-hub 배정 원본. 시트에 붙여넣는 값이 아니라 검증 근거다."""
    ws = wb.create_sheet("ax-hub 배정")
    ws.append(["성함", "역할", "고객사", "교육명", "등급", "일수", "시작일", "종료일",
               "시간(점심 공제)", "적용 시급", "계산액"])
    for a in sorted(assigns, key=lambda x: (x["person"], x["d1"])):
        ws.append([a["person"], a.get("role", ""), a["client"], a["title"], a["qual"], a["days"],
                   dt.datetime.strptime(a["d1"], "%Y-%m-%d"),
                   dt.datetime.strptime(a["d2"], "%Y-%m-%d"),
                   a["hours"], a["rate"], a["amount"]])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=ROLE_FILL.get(ws.cell(r, 2).value, "FFFFFF"))
        ws.cell(r, 7).number_format = "yyyy-mm-dd"
        ws.cell(r, 8).number_format = "yyyy-mm-dd"
        ws.cell(r, 10).number_format = "#,##0"
        ws.cell(r, 11).number_format = "#,##0"
        if ws.cell(r, 11).value is None:
            ws.cell(r, 11).value = "단가미정"
    _style_header(ws, [10, 10, 20, 34, 11, 7, 12, 12, 14, 11, 13])


def _sheet_check(wb, checks):
    """로우데이터 ↔ ax-hub 를 교육 건별로 대조한 결과."""
    ws = wb.create_sheet("검증")
    ws.append(["성함", "판정", "역할(제출)", "역할(ax-hub)", "기간", "로우데이터 행",
               "로우데이터 내용", "제출액", "ax-hub 배정", "계산액", "차액"])
    for c in checks:
        ws.append([c["person"], c["verdict"], c["raw_role"], c["ax_role"], c["period"],
                   c["raw_rows"], c["raw_desc"], c["raw_amount"], c["ax_desc"],
                   c["ax_amount"], c["diff"]])
    for r, c in zip(range(2, ws.max_row + 1), checks):
        fill = VERDICT_FILL.get(ws.cell(r, 2).value)
        if fill:
            ws.cell(r, 2).fill = PatternFill("solid", fgColor=fill)
        if not c["role_ok"]:  # 제출 역할과 배정 역할이 겹치지 않음 — 단가가 섞였을 수 있다
            for col in (3, 4):
                ws.cell(r, col).fill = PatternFill("solid", fgColor="F8CBAD")
        for col in (8, 10, 11):
            ws.cell(r, col).number_format = "#,##0;-#,##0"
    _style_header(ws, [10, 12, 12, 12, 24, 13, 40, 13, 46, 13, 12])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True)
    ap.add_argument("--pm", default="송찬호")
    ap.add_argument("--xlsx", default=br.DEFAULT_XLSX)
    ap.add_argument("--axhub")
    ap.add_argument("--no-rate-memo", action="store_true",
                    help="T(참고/메모)에 ax-hub 기준 시급×시간을 적지 않는다")
    ap.add_argument("--out", required=True)
    a = ap.parse_args()

    sheets = br.load_sheets(a.xlsx)
    rows, warns = br.build(sheets, a.month, a.pm)
    assigns, checks = [], []
    if a.axhub:
        warns += br.crosscheck(rows, a.axhub)
        assigns = br.load_axhub(a.axhub)
        checks = br.reconcile(rows, assigns)
        if not a.no_rate_memo:
            br.apply_rate_memo(rows, assigns)

    header_row = sheets[SHEET].get(1, {})
    headers = [header_row.get(c, "") for c in br.COLS]

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="D9D9D9")
    for i in range(1, len(br.COLS) + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, size=9)
        c.fill = head_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 58

    for r in rows:
        ws.append([r["cells"][c] for c in br.COLS])

    for idx, col in enumerate(br.COLS, 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = WIDTHS.get(col, 12)
        for row_i in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_i, column=idx)
            v = cell.value
            if not v:
                continue
            if col in DATE_COLS:
                cell.value = dt.datetime.strptime(v, "%Y-%m-%d")
                cell.number_format = "yyyy-mm-dd"
            elif col in NUM_COLS:
                cell.value = int(v)
                cell.number_format = "#,##0"
            elif col == "Q":
                cell.alignment = Alignment(wrap_text=False, vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(br.COLS)), ws.max_row)

    if assigns:
        _sheet_axhub(wb, assigns)
    if checks:
        _sheet_check(wb, checks)

    # 확인필요는 붙여넣기 대상이 아니므로 별도 탭으로 뺀다
    ws2 = wb.create_sheet("확인필요")
    ws2.append(["%s / 담당 PM %s / %d행 / 합계 %s원" % (
        a.month, a.pm or "(전체)", len(rows), f"{sum(r['amount'] for r in rows):,}")])
    ws2.append([])
    for w in warns or ["(없음)"]:
        ws2.append([w])
    ws2.column_dimensions["A"].width = 120
    ws2["A1"].font = Font(bold=True)

    out = os.path.abspath(a.out)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("저장: %s (%d행 + 확인필요 %d건)" % (out, len(rows), len(warns)))


if __name__ == "__main__":
    main()
