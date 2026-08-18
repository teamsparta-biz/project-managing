#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ax-hub 배정만으로 인건비 표를 만든다 (로우데이터 제출값과 대조하지 않음).

로우데이터가 아직 안 모였거나, 배정 기준으로 먼저 규모를 보고 싶을 때 쓴다.
지급액은 rates.json 시급 × 시간(점심 공제 반영)이므로 **실제 정산액이 아니라 산정치**다.

usage:
  python axhub_only.py --axhub axhub.json --out "산출물/인건비/2026-07_ax-hub_인건비.xlsx"
"""
import argparse
import datetime as dt
import io
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import build_rows as br

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", newline="\n")

HEADERS = ["성명", "근무시작일", "근무종료일", "고객사명", "교육명",
           "시급", "교육시간\n(점심 공제)", "강의료\n(=시급×시간)",
           "교통비", "숙박비", "세팅", "여비 합계", "총 지급액", "강사/기술튜터"]
WIDTHS = [10, 13, 13, 20, 52, 12, 12, 14, 11, 11, 11, 12, 14, 14]
ROLE_FILL = {"주강사": "DDEBF7", "기술튜터": "FFF2CC"}
# 연노랑 = 사람이 채우거나 고치는 칸.
# 여비는 ax-hub·로우데이터 어디에도 없어 빈 채로 낸다.
# 시급·교육시간은 rates.json·ax-hub 기준값을 미리 넣어두지만 실제 계약과 다를 수 있어 수정 대상이다.
INPUT_FILL = PatternFill("solid", fgColor="FFF9E6")
INPUT_COLS = (6, 7, 9, 10, 11)      # 시급·교육시간·교통비·숙박비·세팅
MONEY_COLS = (6, 8, 9, 10, 11, 12, 13)
HOURS_COL = 7
ROLE_COL = 14


def course_label(client, title):
    """'[고객사명] 교육명'. 교육명 앞에 이미 같은 고객사가 대괄호로 붙어 있으면 겹치지 않게 뗀다."""
    body = title
    m = re.match(r"^\s*\[([^\]]+)\]\s*(.*)$", title)
    if m and br.norm_client(m.group(1)) == br.norm_client(client):
        body = m.group(2)
    return "[%s] %s" % (client, body)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--axhub", required=True)
    ap.add_argument("--out", required=True)
    a = ap.parse_args()

    assigns = sorted(br.load_axhub(a.axhub), key=lambda x: (x["person"], x["d1"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "ax-hub 인건비"
    ws.append(HEADERS)

    total = 0
    print("\t".join(HEADERS))
    for i, x in enumerate(assigns):
        course = course_label(x["client"], x["title"])
        amount = x["amount"]
        total += amount or 0
        r = i + 2
        ws.append([x["person"],
                   dt.datetime.strptime(x["d1"], "%Y-%m-%d"),
                   dt.datetime.strptime(x["d2"], "%Y-%m-%d"),
                   x["client"], course,
                   x.get("rate"),                    # 시급 — 단가미정이면 빈칸
                   x.get("hours"),                   # 교육시간(점심 공제)
                   "=F%d*G%d" % (r, r),              # 강의료 = 시급 × 시간
                   None, None, None,                 # 교통비·숙박비·세팅 — 사람이 입력
                   "=SUM(I%d:K%d)" % (r, r),         # 여비 합계
                   "=H%d+L%d" % (r, r),              # 총 지급액 = 강의료 + 여비
                   x.get("role", "")])
        print("\t".join([x["person"], x["d1"], x["d2"], x["client"], course,
                         "" if x.get("rate") is None else str(x["rate"]),
                         "" if x.get("hours") is None else str(x["hours"]),
                         "" if amount is None else str(amount), "", "", "", "",
                         "" if amount is None else str(amount), x.get("role", "")]))

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).number_format = "yyyy-mm-dd"
        ws.cell(r, 3).number_format = "yyyy-mm-dd"
        for col in MONEY_COLS:
            ws.cell(r, col).number_format = "#,##0"
        ws.cell(r, HOURS_COL).number_format = "0.#"
        for col in INPUT_COLS:
            ws.cell(r, col).fill = INPUT_FILL
        # 시급이 없으면(rates.json null) 0원으로 오해하지 않게 회색으로 덮는다.
        if ws.cell(r, 6).value is None:
            ws.cell(r, 6).fill = PatternFill("solid", fgColor="D9D9D9")
        ws.cell(r, ROLE_COL).fill = PatternFill(
            "solid", fgColor=ROLE_FILL.get(ws.cell(r, ROLE_COL).value, "FFFFFF"))

    fill = PatternFill("solid", fgColor="D9D9D9")
    for i in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, size=9)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(HEADERS)), ws.max_row)

    # 합계 행은 필터 범위 밖에 둔다 (필터를 걸면 같이 딸려 올라간다)
    last = ws.max_row
    tr = last + 2
    ws.cell(tr, 5).value = "합계"
    ws.cell(tr, 5).font = Font(bold=True)
    # 시급은 합계가 의미 없으므로 제외하고, 교육시간은 총 시간이 유용해 더한다.
    for col in [c for c in MONEY_COLS if c != 6] + [HOURS_COL]:
        c = ws.cell(tr, col)
        letter = get_column_letter(col)
        c.value = "=SUM(%s2:%s%d)" % (letter, letter, last)
        c.number_format = "0.#" if col == HOURS_COL else "#,##0"
        c.font = Font(bold=True)

    out = os.path.abspath(a.out)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("# 저장: %s (%d행 / 합계 %s원)" % (out, len(assigns), f"{total:,}"))


if __name__ == "__main__":
    main()
